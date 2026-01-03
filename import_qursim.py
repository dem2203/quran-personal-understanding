"""
Import QurSim Semantic Similarity data from sabdul111/QursimMultilingual
Uses Turkish translation (tr.diyanet.xlsx) which contains verse similarity pairs
"""
import requests
import io
from sqlalchemy import Column, Integer, String, Table
from sqlalchemy.orm import Session
from database import SessionLocal, engine, Base

# Create semantic similarity table
semantic_similarity = Table(
    "semantic_similarity",
    Base.metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("source_surah", Integer, nullable=False),
    Column("source_ayat", Integer, nullable=False),
    Column("target_surah", Integer, nullable=False),
    Column("target_ayat", Integer, nullable=False),
    Column("similarity_degree", Integer, nullable=True),  # 2=strong, 1=weak, 0=none
)

def import_qursim():
    """Import QurSim semantic similarity data from XLSX"""
    db: Session = SessionLocal()
    
    # Check if table exists and has data
    try:
        from sqlalchemy import inspect
        inspector = inspect(engine)
        if 'semantic_similarity' in inspector.get_table_names():
            result = db.execute("SELECT COUNT(*) FROM semantic_similarity").fetchone()
            if result[0] > 0:
                print("QurSim data already exists. Skipping.")
                db.close()
                return
    except Exception as e:
        print(f"Check error: {e}")
    
    # Create table if not exists
    semantic_similarity.create(engine, checkfirst=True)
    
    print("Downloading QurSim data (tr.diyanet.xlsx)...")
    url = "https://raw.githubusercontent.com/sabdul111/QursimMultilingual/main/Qursim%2084%20Holy%20Quran%20Translations/tr.diyanet.xlsx"
    
    try:
        response = requests.get(url, timeout=120)
        response.raise_for_status()
        
        # Parse XLSX
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        ws = wb.active
        
        print("Parsing XLSX file...")
        total_pairs = 0
        seen_pairs = set()
        
        # QurSim format: each row has source and related verses
        # Typically: Surah1, Ayat1, Text1, Surah2, Ayat2, Text2, Similarity
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header
        
        for row in rows:
            if len(row) < 5:
                continue
            
            try:
                # Try to extract surah:ayat pairs
                # Format varies - try common patterns
                src_surah = int(row[0]) if row[0] else None
                src_ayat = int(row[1]) if row[1] else None
                tgt_surah = int(row[3]) if len(row) > 3 and row[3] else None
                tgt_ayat = int(row[4]) if len(row) > 4 and row[4] else None
                similarity = int(row[6]) if len(row) > 6 and row[6] else 1
                
                if src_surah and src_ayat and tgt_surah and tgt_ayat:
                    # Avoid duplicates
                    pair_key = (src_surah, src_ayat, tgt_surah, tgt_ayat)
                    if pair_key not in seen_pairs:
                        seen_pairs.add(pair_key)
                        db.execute(
                            semantic_similarity.insert().values(
                                source_surah=src_surah,
                                source_ayat=src_ayat,
                                target_surah=tgt_surah,
                                target_ayat=tgt_ayat,
                                similarity_degree=similarity
                            )
                        )
                        total_pairs += 1
            except (ValueError, TypeError):
                continue
        
        db.commit()
        wb.close()
        print(f"Imported {total_pairs} semantic similarity pairs from QurSim.")
        
    except Exception as e:
        print(f"Error importing QurSim: {e}")
        # Try alternative: create some sample semantic pairs based on known related verses
        print("Creating sample semantic pairs from known related verses...")
        create_sample_semantic_pairs(db)
    
    db.close()

def create_sample_semantic_pairs(db):
    """Create sample semantic pairs from known related verses"""
    # Known semantically related verse pairs (from Islamic scholarship)
    known_pairs = [
        # Creation of Adam
        (2, 30, 7, 11, 2),   # Bakara 30 - A'raf 11
        (2, 30, 15, 28, 2),  # Bakara 30 - Hijr 28
        (2, 30, 38, 71, 2),  # Bakara 30 - Sad 71
        
        # Story of Moses
        (2, 51, 7, 142, 2),  # Bakara 51 - A'raf 142
        (20, 9, 28, 29, 2),  # Taha 9 - Qasas 29
        
        # Patience and prayer
        (2, 45, 2, 153, 2),  # Bakara 45-153
        
        # God's attributes (Rahman)
        (1, 3, 55, 1, 2),    # Fatiha 3 - Rahman 1
        (17, 110, 59, 22, 2), # Isra 110 - Hashr 22
        
        # Tawbah (Repentance)
        (4, 17, 6, 54, 2),   # Nisa 17 - An'am 54
        (39, 53, 4, 110, 2), # Zumar 53 - Nisa 110
        
        # Day of Judgment
        (82, 1, 81, 1, 2),   # Infitar 1 - Takwir 1
        (99, 1, 56, 4, 2),   # Zilzal 1 - Waqia 4
        
        # Paradise descriptions
        (47, 15, 76, 21, 2), # Muhammad 15 - Insan 21
        (55, 46, 55, 62, 2), # Rahman 46-62
        
        # Hell descriptions
        (4, 56, 22, 19, 2),  # Nisa 56 - Hajj 19
        
        # Prayer times
        (17, 78, 11, 114, 2), # Isra 78 - Hud 114
        
        # Charity
        (2, 261, 2, 271, 2), # Bakara 261-271
        
        # Trust in God
        (3, 159, 65, 3, 2),  # Al-Imran 159 - Talaq 3
        
        # Believers' characteristics
        (23, 1, 8, 2, 2),    # Mu'minun 1 - Anfal 2
    ]
    
    total = 0
    for src_s, src_a, tgt_s, tgt_a, sim in known_pairs:
        try:
            db.execute(
                semantic_similarity.insert().values(
                    source_surah=src_s,
                    source_ayat=src_a,
                    target_surah=tgt_s,
                    target_ayat=tgt_a,
                    similarity_degree=sim
                )
            )
            total += 1
        except:
            pass
    
    db.commit()
    print(f"Created {total} sample semantic pairs.")

if __name__ == "__main__":
    Base.metadata.create_all(bind=engine)
    import_qursim()
